from __future__ import annotations

import importlib.util
import argparse
import subprocess
import sys
import tempfile
import unittest
import uuid
from pathlib import Path

import openpyxl


REPO_ROOT = Path(__file__).resolve().parents[1]
SCRIPT_PATH = REPO_ROOT / "skills" / "MM" / "sap-mb5b-export" / "scripts" / "mb5b_export.py"
SPEC = importlib.util.spec_from_file_location("mb5b_export_under_test", SCRIPT_PATH)
assert SPEC and SPEC.loader
exporter = importlib.util.module_from_spec(SPEC)
sys.modules[SPEC.name] = exporter
SPEC.loader.exec_module(exporter)


def test_directory() -> Path:
    path = Path(tempfile.gettempdir()) / f"sap-mb5b-export-test-{uuid.uuid4().hex}"
    path.mkdir(parents=True)
    return path


def scaled_rect(left: int, top: int, right: int, bottom: int, scale: float):
    return exporter.RectInfo(
        int(left * scale),
        int(top * scale),
        int(right * scale),
        int(bottom * scale),
    )


class WorkbookTests(unittest.TestCase):
    def create_input(self, directory: Path) -> Path:
        path = directory / "targets.xlsx"
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Sheet1"
        worksheet.append(["プラント", "保管場所"])
        worksheet.append(["1002", "117G"])
        worksheet.append(["1002", "105G"])
        workbook.save(path)
        workbook.close()
        return path

    def create_export(self, path: Path, marker: str) -> None:
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.title = "Data"
        worksheet.append(["Plant", "Material", "Batch", "Quantity"])
        worksheet.append(["1002", marker, "B1", 10])
        workbook.save(path)
        workbook.close()

    def test_dates_and_output_names(self) -> None:
        parsed = exporter.parse_export_date("2026-02-28")
        self.assertEqual(parsed.sap_value, "2026/02/28")
        self.assertEqual(parsed.token, "20260228")
        dotted = exporter.format_export_date_for_sap(parsed, "dot")
        self.assertEqual(dotted.sap_value, "2026.02.28")
        self.assertEqual(dotted.token, "20260228")
        path = exporter.individual_output_path(Path("C:/out"), "1002", "117G", parsed.token)
        self.assertEqual(path.name, "MB5B_1002_117G_20260228.xlsx")

    def test_read_targets_and_limit(self) -> None:
        directory = test_directory()
        path = self.create_input(directory)
        targets = exporter.read_targets(path, limit=1)
        self.assertEqual(targets, [{"plant": "1002", "storage": "117G"}])

    def test_storage_column_is_idempotent(self) -> None:
        directory = test_directory()
        path = directory / "individual.xlsx"
        self.create_export(path, "M1")
        exporter.add_storage_column(path, "117G", "保管場所")
        exporter.add_storage_column(path, "117G", "保管場所")
        workbook = openpyxl.load_workbook(path, data_only=True)
        worksheet = workbook["Data"]
        self.assertEqual(worksheet.max_column, 5)
        self.assertEqual(worksheet["D1"].value, "保管場所")
        self.assertEqual(worksheet["D2"].value, "117G")
        workbook.close()

    def test_merge_has_one_header_and_all_rows(self) -> None:
        directory = test_directory()
        first = directory / "first.xlsx"
        second = directory / "second.xlsx"
        self.create_export(first, "M1")
        self.create_export(second, "M2")
        exporter.add_storage_column(first, "117G", "保管場所")
        exporter.add_storage_column(second, "105G", "保管場所")
        output, rows = exporter.merge_plant_files(
            directory,
            "1002",
            [first, second],
            "20260228",
            True,
            None,
        )
        self.assertEqual(rows, 2)
        workbook = openpyxl.load_workbook(output, data_only=True)
        worksheet = workbook["Data"]
        self.assertEqual(worksheet.max_row, 3)
        self.assertEqual(worksheet["D1"].value, "保管場所")
        self.assertEqual([worksheet["D2"].value, worksheet["D3"].value], ["117G", "105G"])
        workbook.close()

    def test_dry_run_command(self) -> None:
        directory = test_directory()
        input_path = self.create_input(directory)
        completed = subprocess.run(
            [
                sys.executable,
                str(SCRIPT_PATH),
                "--input",
                str(input_path),
                "--date",
                "2026-02-28",
                "--output-dir",
                str(directory),
                "--dry-run",
            ],
            check=False,
            capture_output=True,
            text=True,
            encoding="utf-8",
        )
        self.assertEqual(completed.returncode, 0, completed.stderr)
        self.assertIn("MB5B_1002_117G_20260228.xlsx", completed.stdout)
        self.assertIn("MB5B_1002_20260228.xlsx", completed.stdout)

    def test_partial_failure_returns_two_and_merges_only_successes(self) -> None:
        directory = test_directory()
        input_path = self.create_input(directory)
        args = argparse.Namespace(
            input=input_path,
            date=exporter.parse_export_date("2026-02-28"),
            sap_date_format="slash",
            output_dir=directory,
            sheet="Sheet1",
            plant_column="プラント",
            storage_column="保管場所",
            storage_header="保管場所",
            limit=0,
            dry_run=False,
            overwrite=True,
        )

        original_connect = exporter.connect_sap_with_access_helper
        original_run = exporter.run_one_sap_export
        original_require = exporter.require_windows_runtime

        def fake_run(_session, config, _plant, storage, output_path, _log_file):
            if storage == "105G":
                raise RuntimeError("synthetic export failure")
            self.create_export(output_path, "M1")

        try:
            exporter.require_windows_runtime = lambda: None
            exporter.connect_sap_with_access_helper = lambda _directory: object()
            exporter.run_one_sap_export = fake_run
            result = exporter.run_main(args)
        finally:
            exporter.connect_sap_with_access_helper = original_connect
            exporter.run_one_sap_export = original_run
            exporter.require_windows_runtime = original_require

        self.assertEqual(result, 2)
        merged = directory / "MB5B_1002_20260228.xlsx"
        workbook = openpyxl.load_workbook(merged, data_only=True)
        worksheet = workbook["Data"]
        self.assertEqual(worksheet.max_row, 2)
        self.assertEqual(worksheet["D2"].value, "117G")
        workbook.close()


class SapSelectionScreenTests(unittest.TestCase):
    class FakeControl:
        def __init__(self, selected: bool) -> None:
            self.Selected = selected

    class FakeSession:
        def __init__(self) -> None:
            self.controls = {
                exporter.SAP_TOTALS_HIERARCHICAL_ID: SapSelectionScreenTests.FakeControl(True),
                exporter.SAP_TOTALS_NON_HIERARCHICAL_ID: SapSelectionScreenTests.FakeControl(False),
            }

        def FindById(self, control_id: str):
            return self.controls[control_id]

    def test_totals_non_hierarchical_scope_is_enforced(self) -> None:
        session = self.FakeSession()
        selected_by = exporter.select_totals_non_hierarchical_representation(session)
        self.assertIn(exporter.SAP_TOTALS_NON_HIERARCHICAL_ID, selected_by)
        self.assertFalse(session.controls[exporter.SAP_TOTALS_HIERARCHICAL_ID].Selected)
        self.assertTrue(session.controls[exporter.SAP_TOTALS_NON_HIERARCHICAL_ID].Selected)

    def test_excel_export_format_uses_control_id(self) -> None:
        window = exporter.WindowInfo(
            1,
            100,
            exporter.SAP_FRONTEND_SESSION_CLASS,
            exporter.RectInfo(0, 0, 800, 400),
            (
                exporter.ControlInfo(0, "button", "Button", "", 113, exporter.RectInfo(10, 10, 50, 30), label="任意"),
                exporter.ControlInfo(
                    1,
                    "button",
                    "Button",
                    "",
                    exporter.SAP_EXCEL_EXPORT_FORMAT_CONTROL_ID,
                    exporter.RectInfo(60, 10, 100, 30),
                    label="任意",
                ),
            ),
        )
        self.assertEqual(exporter.select_excel_export_format_button(window), 1)

    def test_export_to_dialog_uses_control_id_pair(self) -> None:
        window = exporter.WindowInfo(
            2,
            100,
            exporter.COMMON_DIALOG_CLASS,
            exporter.RectInfo(0, 0, 400, 160),
            (
                exporter.ControlInfo(
                    0,
                    "button",
                    "Button",
                    "",
                    exporter.SAP_EXPORT_TO_CONTROL_ID,
                    exporter.RectInfo(200, 100, 260, 130),
                    label="任意",
                ),
                exporter.ControlInfo(
                    1,
                    "button",
                    "Button",
                    "",
                    exporter.SAP_EXPORT_CANCEL_CONTROL_ID,
                    exporter.RectInfo(270, 100, 330, 130),
                    label="任意",
                ),
            ),
        )
        self.assertEqual(exporter.select_export_to_button(window), 0)

    def test_classical_list_table_is_extracted_from_label_ids(self) -> None:
        labels = {
            (1, 1): "Plnt",
            (6, 1): "Material",
            (47, 1): "From Date",
            (63, 1): "To Date",
            (79, 1): "Opening Stock",
            (1, 3): "1710",
            (6, 3): "FG126",
            (47, 3): "0000.01.01",
            (63, 3): "2020.12.31",
            (79, 3): "0",
            (1, 4): "1710",
            (6, 4): "FG129",
            (47, 4): "0000.01.01",
            (63, 4): "2020.12.31",
            (79, 4): "5,300",
        }
        headers, rows = exporter.extract_table_from_sap_labels(labels)
        self.assertEqual(headers, ["Plnt", "Material", "From Date", "To Date", "Opening Stock"])
        self.assertEqual(
            rows,
            [
                ["1710", "FG126", "0000.01.01", "2020.12.31", "0"],
                ["1710", "FG129", "0000.01.01", "2020.12.31", "5,300"],
            ],
        )


class LanguageIndependentSelectorTests(unittest.TestCase):
    LABEL_SETS = (
        ("Export", "Destination", "Cancel"),
        ("エクスポート", "エクスポート先", "中止"),
        ("导出", "导出位置", "取消"),
        ("匯出", "匯出位置", "取消"),
        ("Δοκιμή", "保存先", "Отмена"),
    )

    def export_window(self, labels: tuple[str, str, str], scale: float):
        controls = (
            exporter.ControlInfo(0, "edit", "Edit", "", None, scaled_rect(200, 100, 800, 130, scale), label=labels[0]),
            exporter.ControlInfo(1, "combobox", "ComboBox", "", None, scaled_rect(200, 140, 650, 170, scale)),
            exporter.ControlInfo(2, "combobox", "ComboBox", "", None, scaled_rect(200, 180, 650, 210, scale)),
            exporter.ControlInfo(3, "button", "Button", "", None, scaled_rect(710, 260, 820, 290, scale), label=labels[1]),
            exporter.ControlInfo(4, "button", "Button", "", None, scaled_rect(830, 260, 900, 290, scale), label=labels[2]),
        )
        return exporter.WindowInfo(
            1,
            100,
            "Chrome_WidgetWin_1",
            scaled_rect(0, 0, 920, 320, scale),
            controls,
            labels[0],
        )

    def test_export_selection_ignores_language_and_scale(self) -> None:
        for scale in (1.0, 1.25, 1.5):
            for labels in self.LABEL_SETS:
                with self.subTest(scale=scale, labels=labels):
                    window = self.export_window(labels, scale)
                    self.assertGreater(exporter.score_export_dialog(window), 0)
                    self.assertEqual(exporter.select_export_filename_control(window), 0)
                    self.assertEqual(exporter.select_export_destination_button(window), 3)

    def test_standard_save_ids_ignore_labels(self) -> None:
        for labels in self.LABEL_SETS:
            controls = (
                exporter.ControlInfo(0, "edit", "Edit", "", 0x480, exporter.RectInfo(20, 350, 600, 380), label=labels[0]),
                exporter.ControlInfo(1, "button", "Button", "", 1, exporter.RectInfo(620, 350, 700, 380), label=labels[1]),
                exporter.ControlInfo(2, "button", "Button", "", 2, exporter.RectInfo(710, 350, 790, 380), label=labels[2]),
            )
            window = exporter.WindowInfo(2, 100, "#32770", exporter.RectInfo(0, 0, 800, 420), controls)
            self.assertGreater(exporter.score_save_dialog(window), 0)
            self.assertEqual(exporter.select_save_filename_control(window), 0)
            self.assertEqual(exporter.select_standard_button(window, 1), 1)

    def test_overwrite_requires_standard_yes_id(self) -> None:
        controls = (
            exporter.ControlInfo(0, "button", "Button", "", 6, exporter.RectInfo(500, 200, 580, 230), label="任意"),
            exporter.ControlInfo(1, "button", "Button", "", 7, exporter.RectInfo(590, 200, 670, 230), label="文字"),
        )
        window = exporter.WindowInfo(3, 100, "#32770", exporter.RectInfo(0, 0, 700, 260), controls)
        self.assertEqual(exporter.select_standard_button(window, 6), 0)


if __name__ == "__main__":
    unittest.main()
