from __future__ import annotations

import argparse
import json
import os
import platform
import subprocess
import sys
import time
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path, PureWindowsPath
from typing import Any, Iterable, Sequence

import openpyxl

WINDOWS_IMPORT_ERROR: Exception | None = None
try:
    import pythoncom
    import win32clipboard
    import win32com.client
    from pywinauto import Desktop, keyboard, mouse
except Exception as exc:  # pragma: no cover - exercised by preflight on non-Windows hosts
    WINDOWS_IMPORT_ERROR = exc
    pythoncom = None
    win32clipboard = None
    win32com = None
    Desktop = None
    keyboard = None
    mouse = None


FILE_PREFIX = "MB5B"
DATA_SHEET = "Data"
DEFAULT_INPUT_SHEET = "Sheet1"
DEFAULT_PLANT_HEADER = "プラント"
DEFAULT_STORAGE_HEADER = "保管場所"
SAP_EXPORT_BUTTON_ID = "wnd[0]/tbar[1]/btn[43]"
SAP_EXECUTE_BUTTON_ID = "wnd[0]/tbar[1]/btn[8]"
SAP_PLANT_FIELD_ID = "wnd[0]/usr/ctxtWERKS-LOW"
SAP_STORAGE_FIELD_ID = "wnd[0]/usr/ctxtLGORT-LOW"
SAP_DATE_FIELD_ID = "wnd[0]/usr/ctxtDATUM-HIGH"
SAP_STOCK_TYPE_ID = "wnd[0]/usr/radLGBST"
SAP_EXPORT_AS_FILE_ID = (
    "wnd[1]/usr/ssubSUB_CONFIGURATION:SAPLSALV_GUI_CUL_EXPORT_AS:0512/"
    "txtGS_EXPORT-FILE_NAME"
)
SAP_EXPORT_AS_BUTTON_ID = "wnd[1]/tbar[0]/btn[20]"
SAP_SAVE_FILE_PATH_ID = "wnd[1]/usr/ctxtDY_PATH"
SAP_SAVE_FILE_NAME_ID = "wnd[1]/usr/ctxtDY_FILENAME"
SAP_SAVE_FILE_REPLACE_BUTTON_ID = "wnd[1]/tbar[0]/btn[11]"
SAP_SAVE_FILE_GENERATE_BUTTON_ID = "wnd[1]/tbar[0]/btn[0]"

COMMON_DIALOG_CLASS = "#32770"
STANDARD_FILENAME_CONTROL_ID = 0x480
STANDARD_OK_CONTROL_ID = 1
STANDARD_YES_CONTROL_ID = 6
KNOWN_FILENAME_AUTOMATION_IDS = {
    "1001",
    "1148",
    "filenamecontrolhost",
    "filenamecontrol",
}
KNOWN_OK_AUTOMATION_IDS = {"1", "filesave", "savebutton"}
SAP_SCRIPTING_SECURITY_PROMPTS = (
    "A script is attempting to access SAP GUI.",
    "A script is opening a connection to system:",
)


@dataclass(frozen=True)
class ExportDate:
    sap_value: str
    token: str


@dataclass(frozen=True)
class RunConfig:
    input_path: Path
    export_date: ExportDate
    output_dir: Path
    sheet: str
    plant_column: str
    storage_column: str
    storage_header: str
    limit: int
    dry_run: bool
    overwrite: bool


@dataclass(frozen=True)
class RectInfo:
    left: int
    top: int
    right: int
    bottom: int

    @property
    def width(self) -> int:
        return max(0, self.right - self.left)

    @property
    def height(self) -> int:
        return max(0, self.bottom - self.top)


@dataclass(frozen=True)
class ControlInfo:
    index: int
    control_type: str
    class_name: str
    automation_id: str
    control_id: int | None
    rect: RectInfo
    enabled: bool = True
    visible: bool = True
    label: str = ""


@dataclass(frozen=True)
class WindowInfo:
    handle: int
    process_id: int
    class_name: str
    rect: RectInfo
    controls: tuple[ControlInfo, ...]
    title: str = ""
    owner_handle: int = 0


def log_line(log_file: Path | None, message: str) -> None:
    line = f"{datetime.now():%Y/%m/%d %H:%M:%S} {message}"
    print(line, flush=True)
    if log_file:
        log_file.parent.mkdir(parents=True, exist_ok=True)
        with log_file.open("a", encoding="utf-8") as stream:
            stream.write(line + "\n")


def parse_export_date(value: str) -> ExportDate:
    normalized = value.strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            parsed = datetime.strptime(normalized, fmt)
            return ExportDate(parsed.strftime("%Y.%m.%d"), parsed.strftime("%Y%m%d"))
        except ValueError:
            continue
    raise argparse.ArgumentTypeError(
        f"invalid date {value!r}; use YYYY-MM-DD, YYYY/MM/DD, YYYY.MM.DD, or YYYYMMDD"
    )


def safe_filename_part(value: str) -> str:
    result = str(value).strip()
    for character in '\\/:*?"<>|':
        result = result.replace(character, "_")
    return result


def cell_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def read_targets(
    input_path: Path,
    sheet: str = DEFAULT_INPUT_SHEET,
    plant_column: str = DEFAULT_PLANT_HEADER,
    storage_column: str = DEFAULT_STORAGE_HEADER,
    limit: int = 0,
) -> list[dict[str, str]]:
    if not input_path.exists():
        raise FileNotFoundError(f"input workbook not found: {input_path}")

    workbook = openpyxl.load_workbook(input_path, data_only=True, read_only=True)
    try:
        worksheet = workbook[sheet] if sheet in workbook.sheetnames else workbook[workbook.sheetnames[0]]
        headers = {
            cell_text(cell.value): index
            for index, cell in enumerate(worksheet[1], start=1)
            if cell.value is not None
        }
        missing = [header for header in (plant_column, storage_column) if header not in headers]
        if missing:
            raise ValueError(f"required header(s) not found: {', '.join(missing)}")

        plant_index = headers[plant_column] - 1
        storage_index = headers[storage_column] - 1
        targets: list[dict[str, str]] = []
        for row in worksheet.iter_rows(min_row=2, values_only=True):
            plant = cell_text(row[plant_index]) if plant_index < len(row) else ""
            storage = cell_text(row[storage_index]) if storage_index < len(row) else ""
            if not plant or not storage:
                continue
            targets.append({"plant": plant, "storage": storage})
            if limit and len(targets) >= limit:
                break
        return targets
    finally:
        workbook.close()


def individual_output_path(
    output_dir: Path,
    plant: str,
    storage: str,
    date_token: str,
) -> Path:
    return output_dir / (
        f"{FILE_PREFIX}_{safe_filename_part(plant)}_"
        f"{safe_filename_part(storage)}_{date_token}.xlsx"
    )


def plant_merge_output_path(output_dir: Path, plant: str, date_token: str) -> Path:
    return output_dir / f"{FILE_PREFIX}_{safe_filename_part(plant)}_{date_token}.xlsx"


def require_windows_runtime() -> None:
    if os.name != "nt":
        raise RuntimeError("SAP GUI automation requires Windows")
    if WINDOWS_IMPORT_ERROR is not None:
        raise RuntimeError(f"Windows automation dependency unavailable: {WINDOWS_IMPORT_ERROR}")


def get_sap_session():
    require_windows_runtime()
    pythoncom.CoInitialize()
    sap_gui = win32com.client.GetObject("SAPGUI")
    application = sap_gui.GetScriptingEngine
    if application.Children.Count == 0:
        raise RuntimeError("SAP GUI has no active connection")
    connection = application.Children(0)
    if connection.Children.Count == 0:
        raise RuntimeError("SAP GUI has no active session")
    return connection.Children(0)


def wait_sap(session, timeout: int = 300) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            if not session.Busy:
                time.sleep(0.4)
                return
        except Exception:
            time.sleep(0.4)
            return
        time.sleep(0.5)
    raise TimeoutError("SAP did not become ready in time")


def sap_info(session) -> str:
    try:
        return (
            f"transaction={session.Info.Transaction}, program={session.Info.Program}, "
            f"screen={session.Info.ScreenNumber}"
        )
    except Exception as exc:
        return f"session info unavailable: {exc}"


def sap_status_text(session) -> str:
    try:
        return str(session.FindById("wnd[0]/sbar").Text)
    except Exception:
        return ""


def sap_required(session, control_id: str):
    try:
        return session.FindById(control_id)
    except Exception as exc:
        raise RuntimeError(f"missing SAP control: {control_id} ({exc})") from exc


def has_sap_control(session, control_id: str) -> bool:
    try:
        session.FindById(control_id)
        return True
    except Exception:
        return False


def wait_for_mb5b_selection_screen(session, timeout: int = 30) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        wait_sap(session, timeout=5)
        if has_sap_control(session, SAP_PLANT_FIELD_ID) and has_sap_control(
            session, SAP_DATE_FIELD_ID
        ):
            return
        time.sleep(0.5)
    raise TimeoutError(f"MB5B selection screen was not ready: {sap_info(session)}")


def select_storage_location_batch_stock(session) -> str:
    try:
        sap_required(session, SAP_STOCK_TYPE_ID).Select()
        return SAP_STOCK_TYPE_ID
    except Exception:
        pass

    def walk(control) -> bool:
        try:
            children = control.Children
        except Exception:
            return False
        for child in children:
            technical_values: list[str] = []
            for attribute in ("Id", "Name"):
                try:
                    technical_values.append(str(getattr(child, attribute)))
                except Exception:
                    pass
            technical_identity = " ".join(technical_values).upper()
            if "RADLGBST" in technical_identity or technical_identity.endswith(" LGBST"):
                try:
                    child.Select()
                    return True
                except Exception:
                    pass
            if walk(child):
                return True
        return False

    if walk(sap_required(session, "wnd[0]/usr")):
        return "technical-tree-search"
    raise RuntimeError(f"could not select SAP technical control {SAP_STOCK_TYPE_ID}")


def set_clipboard_text(value: str) -> None:
    require_windows_runtime()
    win32clipboard.OpenClipboard()
    try:
        win32clipboard.EmptyClipboard()
        win32clipboard.SetClipboardData(win32clipboard.CF_UNICODETEXT, value)
    finally:
        win32clipboard.CloseClipboard()


def rect_info(rectangle) -> RectInfo:
    return RectInfo(
        int(rectangle.left),
        int(rectangle.top),
        int(rectangle.right),
        int(rectangle.bottom),
    )


def safe_call(obj, method: str, default: Any = None) -> Any:
    try:
        value = getattr(obj, method)
        return value() if callable(value) else value
    except Exception:
        return default


def normalize_control_type(wrapper) -> str:
    element_info = safe_call(wrapper, "element_info")
    control_type = getattr(element_info, "control_type", "") if element_info else ""
    if control_type:
        return str(control_type).lower()
    class_name = str(safe_call(wrapper, "class_name", "")).lower()
    if "button" in class_name:
        return "button"
    if class_name == "edit" or "edit" in class_name:
        return "edit"
    if "combo" in class_name:
        return "combobox"
    return class_name


def control_info(index: int, wrapper) -> ControlInfo | None:
    rectangle = safe_call(wrapper, "rectangle")
    if rectangle is None:
        return None
    raw_control_id = safe_call(wrapper, "control_id")
    try:
        parsed_control_id = int(raw_control_id) if raw_control_id is not None else None
    except (TypeError, ValueError):
        parsed_control_id = None
    return ControlInfo(
        index=index,
        control_type=normalize_control_type(wrapper),
        class_name=str(safe_call(wrapper, "class_name", "")),
        automation_id=str(safe_call(wrapper, "automation_id", "")),
        control_id=parsed_control_id,
        rect=rect_info(rectangle),
        enabled=bool(safe_call(wrapper, "is_enabled", True)),
        visible=bool(safe_call(wrapper, "is_visible", True)),
        label=str(safe_call(wrapper, "window_text", "")),
    )


def window_handle(window) -> int:
    handle = safe_call(window, "handle", 0)
    try:
        return int(handle)
    except (TypeError, ValueError):
        return 0


def enumerate_control_wrappers(window) -> list[Any]:
    handle = window_handle(window)
    wrappers: list[Any] = []
    if handle and Desktop is not None:
        try:
            uia_window = Desktop(backend="uia").window(handle=handle)
            wrappers = list(uia_window.descendants())
        except Exception:
            wrappers = []
    if not wrappers:
        try:
            wrappers = list(window.descendants())
        except Exception:
            try:
                wrappers = list(window.children())
            except Exception:
                wrappers = []
    return wrappers


def snapshot_window(window) -> tuple[WindowInfo, list[Any]] | None:
    rectangle = safe_call(window, "rectangle")
    if rectangle is None:
        return None
    wrappers = enumerate_control_wrappers(window)
    infos: list[ControlInfo] = []
    retained_wrappers: list[Any] = []
    for wrapper in wrappers:
        info = control_info(len(infos), wrapper)
        if info is None:
            continue
        infos.append(info)
        retained_wrappers.append(wrapper)
    process_id = safe_call(window, "process_id", 0)
    try:
        process_id = int(process_id)
    except (TypeError, ValueError):
        process_id = 0
    owner_handle = 0
    try:
        owner_handle = window_handle(window.owner())
    except Exception:
        pass
    return (
        WindowInfo(
            handle=window_handle(window),
            process_id=process_id,
            class_name=str(safe_call(window, "class_name", "")),
            rect=rect_info(rectangle),
            controls=tuple(infos),
            title=str(safe_call(window, "window_text", "")),
            owner_handle=owner_handle,
        ),
        retained_wrappers,
    )


def usable_top_windows() -> list[Any]:
    require_windows_runtime()
    try:
        windows = Desktop(backend="win32").windows()
    except Exception:
        time.sleep(0.2)
        windows = Desktop(backend="win32").windows()
    usable: list[Any] = []
    for window in windows:
        try:
            if not window.is_visible():
                continue
            rectangle = window.rectangle()
            if rectangle.right <= rectangle.left or rectangle.bottom <= rectangle.top:
                continue
            if rectangle.left < -1000 or rectangle.top < -1000:
                continue
            usable.append(window)
        except Exception:
            continue
    return usable


def controls_of_type(window: WindowInfo, *types: str) -> list[ControlInfo]:
    wanted = {value.lower() for value in types}
    return [
        control
        for control in window.controls
        if control.visible and control.enabled and control.control_type.lower() in wanted
    ]


def score_export_dialog(window: WindowInfo) -> int:
    edits = controls_of_type(window, "edit")
    combos = controls_of_type(window, "combobox", "combo box")
    buttons = controls_of_type(window, "button")
    if not edits or len(buttons) < 2:
        return 0
    if window.rect.width < 450 or window.rect.height < 180:
        return 0
    score = 10
    score += min(len(combos), 3) * 8
    score += min(len(buttons), 4) * 4
    if window.class_name != COMMON_DIALOG_CLASS:
        score += 10
    return score


def score_save_dialog(window: WindowInfo) -> int:
    edits = controls_of_type(window, "edit")
    buttons = controls_of_type(window, "button")
    if not edits or not buttons:
        return 0
    score = 0
    if window.class_name == COMMON_DIALOG_CLASS:
        score += 30
    if any(control.control_id == STANDARD_FILENAME_CONTROL_ID for control in edits):
        score += 60
    if any(
        control.automation_id.lower() in KNOWN_FILENAME_AUTOMATION_IDS for control in edits
    ):
        score += 50
    if any(control.control_id == STANDARD_OK_CONTROL_ID for control in buttons):
        score += 20
    if window.rect.width >= 500 and window.rect.height >= 300:
        score += 10
    return score


def select_export_filename_control(window: WindowInfo) -> int | None:
    edits = controls_of_type(window, "edit")
    if not edits:
        return None
    upper_limit = window.rect.top + int(window.rect.height * 0.65)
    upper_edits = [control for control in edits if control.rect.top < upper_limit]
    candidates = upper_edits or edits
    candidates.sort(key=lambda control: (control.rect.width, -control.rect.top), reverse=True)
    return candidates[0].index


def select_export_destination_button(window: WindowInfo) -> int | None:
    buttons = controls_of_type(window, "button")
    if not buttons:
        return None
    lower_limit = window.rect.top + int(window.rect.height * 0.55)
    right_limit = window.rect.left + int(window.rect.width * 0.45)
    candidates = [
        control
        for control in buttons
        if control.rect.top >= lower_limit and control.rect.left >= right_limit
    ]
    if not candidates:
        return None
    candidates.sort(key=lambda control: (control.rect.left, control.rect.top))
    return candidates[0].index


def select_save_filename_control(window: WindowInfo) -> int | None:
    edits = controls_of_type(window, "edit")
    for control in edits:
        if control.control_id == STANDARD_FILENAME_CONTROL_ID:
            return control.index
    for control in edits:
        if control.automation_id.lower() in KNOWN_FILENAME_AUTOMATION_IDS:
            return control.index
    if not edits:
        return None
    lower_limit = window.rect.top + int(window.rect.height * 0.45)
    lower_edits = [control for control in edits if control.rect.top >= lower_limit]
    candidates = lower_edits or edits
    candidates.sort(key=lambda control: (control.rect.top, control.rect.width), reverse=True)
    return candidates[0].index


def select_standard_button(window: WindowInfo, control_id: int) -> int | None:
    buttons = controls_of_type(window, "button")
    for control in buttons:
        if control.control_id == control_id:
            return control.index
    if control_id == STANDARD_OK_CONTROL_ID:
        for control in buttons:
            if control.automation_id.lower() in KNOWN_OK_AUTOMATION_IDS:
                return control.index
    return None


def is_single_ok_dialog(window: WindowInfo) -> bool:
    if controls_of_type(window, "edit"):
        return False
    buttons = controls_of_type(window, "button")
    return len(buttons) == 1 and select_standard_button(window, STANDARD_OK_CONTROL_ID) is not None


def is_sap_scripting_security_prompt(window: WindowInfo) -> bool:
    if window.class_name != COMMON_DIALOG_CLASS:
        return False
    labels = [control.label.strip() for control in window.controls if control.label]
    if not any(
        any(prompt in label for prompt in SAP_SCRIPTING_SECURITY_PROMPTS)
        for label in labels
    ):
        return False
    return select_standard_button(window, STANDARD_OK_CONTROL_ID) is not None


def click_sap_scripting_security_prompt(
    window: WindowInfo,
    wrappers: Sequence[Any],
) -> bool:
    button_index = select_standard_button(window, STANDARD_OK_CONTROL_ID)
    if button_index is None:
        return False
    click_wrapper(wrappers[button_index])
    return True


def click_wrapper(wrapper) -> None:
    try:
        wrapper.click_input()
        return
    except Exception:
        pass
    try:
        wrapper.invoke()
        return
    except Exception as exc:
        raise RuntimeError(f"could not activate UI control: {exc}") from exc


def set_wrapper_text(wrapper, value: str) -> None:
    try:
        wrapper.set_edit_text(value)
        return
    except Exception:
        pass
    try:
        wrapper.set_focus()
        set_clipboard_text(value)
        keyboard.send_keys("^a")
        keyboard.send_keys("^v")
        return
    except Exception as exc:
        raise RuntimeError(f"could not set UI edit control: {exc}") from exc


def diagnostic_payload(window: WindowInfo) -> dict[str, Any]:
    return {
        "handle": window.handle,
        "process_id": window.process_id,
        "class_name": window.class_name,
        "title": window.title,
        "owner_handle": window.owner_handle,
        "rect": window.rect.__dict__,
        "controls": [
            {
                "index": control.index,
                "control_type": control.control_type,
                "class_name": control.class_name,
                "automation_id": control.automation_id,
                "control_id": control.control_id,
                "rect": control.rect.__dict__,
                "enabled": control.enabled,
                "visible": control.visible,
                "label": control.label,
            }
            for control in window.controls
        ],
    }


def save_window_diagnostic(
    window,
    window_info: WindowInfo,
    directory: Path,
    prefix: str,
) -> None:
    directory.mkdir(parents=True, exist_ok=True)
    stem = f"{prefix}_{window_info.handle}_{datetime.now():%Y%m%d_%H%M%S_%f}"
    json_path = directory / f"{stem}.json"
    json_path.write_text(
        json.dumps(diagnostic_payload(window_info), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    try:
        window.capture_as_image().save(directory / f"{stem}.png")
    except Exception:
        pass


def inspect_ui(output_dir: Path) -> int:
    require_windows_runtime()
    diagnostic_dir = output_dir / f"MB5B_ui_inspection_{datetime.now():%Y%m%d_%H%M%S}"
    diagnostic_dir.mkdir(parents=True, exist_ok=False)
    payload = {
        "created_at": datetime.now().isoformat(),
        "platform": platform.platform(),
        "python": sys.version,
        "windows": [],
    }
    for window in usable_top_windows():
        snapshot = snapshot_window(window)
        if snapshot is None:
            continue
        info, _ = snapshot
        payload["windows"].append(diagnostic_payload(info))
        try:
            window.capture_as_image().save(diagnostic_dir / f"window_{info.handle}.png")
        except Exception:
            pass
    (diagnostic_dir / "ui-tree.json").write_text(
        json.dumps(payload, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    print(diagnostic_dir)
    return 0


def detect_sap_process_ids() -> set[int]:
    process_ids: set[int] = set()
    for window in usable_top_windows():
        class_name = str(safe_call(window, "class_name", "")).casefold()
        if "sap_frontend" not in class_name and "sapgui" not in class_name:
            continue
        process_id = safe_call(window, "process_id", 0)
        try:
            process_ids.add(int(process_id))
        except (TypeError, ValueError):
            pass
    process_ids.discard(0)
    return process_ids


def find_new_dialogs(
    baseline_handles: set[int],
    allowed_process_ids: set[int] | None = None,
) -> list[tuple[Any, WindowInfo, list[Any]]]:
    candidates: list[tuple[Any, WindowInfo, list[Any]]] = []
    for window in usable_top_windows():
        snapshot = snapshot_window(window)
        if snapshot is None:
            continue
        info, wrappers = snapshot
        if info.handle in baseline_handles:
            continue
        if (
            allowed_process_ids
            and info.process_id not in allowed_process_ids
            and info.owner_handle not in baseline_handles
        ):
            continue
        if info.rect.width < 180 or info.rect.height < 100:
            continue
        candidates.append((window, info, wrappers))
    return candidates


def close_tracked_windows(handles: Iterable[int]) -> None:
    handle_set = set(handles)
    if not handle_set:
        return
    for window in usable_top_windows():
        if window_handle(window) not in handle_set:
            continue
        try:
            window.close()
        except Exception:
            try:
                window.set_focus()
                keyboard.send_keys("{ESC}")
            except Exception:
                pass


def dialog_helper(
    output_path: Path,
    overwrite: bool,
    timeout: int = 300,
    ready_file: Path | None = None,
    allowed_process_ids: set[int] | None = None,
) -> int:
    require_windows_runtime()
    start_time = time.time()
    baseline_handles = {window_handle(window) for window in usable_top_windows()}
    if allowed_process_ids is None:
        allowed_process_ids = detect_sap_process_ids()
    diagnostic_dir = output_path.parent / f"MB5B_diagnostics_{datetime.now():%Y%m%d_%H%M%S}"
    helper_log = diagnostic_dir / "dialog-helper.log"
    state = "wait-export"
    tracked_handles: set[int] = set()
    submitted_states: set[str] = set()
    deadline = start_time + timeout
    log_line(helper_log, f"helper started output={output_path} overwrite={overwrite}")
    if ready_file is not None:
        ready_file.write_text("ready", encoding="ascii")

    try:
        while time.time() < deadline:
            if output_path.exists() and output_path.stat().st_mtime >= start_time - 1:
                log_line(helper_log, "output file detected")
                close_tracked_windows(tracked_handles)
                return 0

            candidates = find_new_dialogs(baseline_handles, allowed_process_ids)
            for window, info, wrappers in candidates:
                tracked_handles.add(info.handle)
                if is_sap_scripting_security_prompt(info):
                    state_key = f"sap-security-{info.handle}"
                    if state_key not in submitted_states:
                        click_sap_scripting_security_prompt(info, wrappers)
                        submitted_states.add(state_key)
                        log_line(
                            helper_log,
                            f"SAP scripting security prompt confirmed handle={info.handle}",
                        )
                    continue

                if state == "wait-export":
                    export_score = score_export_dialog(info)
                    if export_score <= 0:
                        if is_single_ok_dialog(info):
                            button_index = select_standard_button(info, STANDARD_OK_CONTROL_ID)
                            if button_index is not None and "access-ok" not in submitted_states:
                                click_wrapper(wrappers[button_index])
                                submitted_states.add("access-ok")
                        continue
                    save_window_diagnostic(window, info, diagnostic_dir, "export-dialog")
                    edit_index = select_export_filename_control(info)
                    button_index = select_export_destination_button(info)
                    if edit_index is None or button_index is None:
                        log_line(helper_log, "export dialog structure is unsupported")
                        return 10
                    if "export" in submitted_states:
                        continue
                    set_wrapper_text(wrappers[edit_index], output_path.stem)
                    click_wrapper(wrappers[button_index])
                    submitted_states.add("export")
                    state = "wait-save"
                    log_line(helper_log, f"export dialog submitted handle={info.handle}")
                    continue

                if state in {"wait-save", "wait-output"}:
                    save_score = score_save_dialog(info)
                    if save_score > 0:
                        if "save" in submitted_states:
                            continue
                        save_window_diagnostic(window, info, diagnostic_dir, "save-dialog")
                        edit_index = select_save_filename_control(info)
                        button_index = select_standard_button(info, STANDARD_OK_CONTROL_ID)
                        if edit_index is None or button_index is None:
                            log_line(helper_log, "save dialog structure is unsupported")
                            return 11
                        set_wrapper_text(wrappers[edit_index], str(output_path))
                        click_wrapper(wrappers[button_index])
                        submitted_states.add("save")
                        state = "wait-output"
                        log_line(helper_log, f"save dialog submitted handle={info.handle}")
                        continue

                    if is_single_ok_dialog(info):
                        button_index = select_standard_button(info, STANDARD_OK_CONTROL_ID)
                        if button_index is not None and f"ok-{info.handle}" not in submitted_states:
                            save_window_diagnostic(window, info, diagnostic_dir, "single-ok-dialog")
                            click_wrapper(wrappers[button_index])
                            submitted_states.add(f"ok-{info.handle}")
                        continue

                    buttons = controls_of_type(info, "button")
                    edits = controls_of_type(info, "edit")
                    if not edits and len(buttons) >= 2:
                        yes_index = select_standard_button(info, STANDARD_YES_CONTROL_ID)
                        if yes_index is not None and overwrite:
                            if "overwrite" not in submitted_states:
                                save_window_diagnostic(window, info, diagnostic_dir, "overwrite-dialog")
                                click_wrapper(wrappers[yes_index])
                                submitted_states.add("overwrite")
                                log_line(helper_log, "overwrite confirmation submitted")
                            continue
                        save_window_diagnostic(window, info, diagnostic_dir, "ambiguous-dialog")
                        log_line(
                            helper_log,
                            "ambiguous multi-button dialog detected; refusing language-dependent click",
                        )
                        return 12

            time.sleep(0.35)
    except Exception as exc:
        log_line(helper_log, f"helper error: {type(exc).__name__}: {exc}")
        return 13
    finally:
        if not output_path.exists() or output_path.stat().st_mtime < start_time - 1:
            close_tracked_windows(tracked_handles)

    log_line(helper_log, f"helper timeout state={state}")
    return 14


def access_helper(timeout: int = 120, ready_file: Path | None = None) -> int:
    require_windows_runtime()
    baseline_handles = {window_handle(window) for window in usable_top_windows()}
    allowed_process_ids = detect_sap_process_ids()
    if ready_file is not None:
        ready_file.write_text("ready", encoding="ascii")
    deadline = time.time() + timeout
    while time.time() < deadline:
        for _window, info, wrappers in find_new_dialogs(
            baseline_handles, allowed_process_ids
        ):
            if is_sap_scripting_security_prompt(info):
                if click_sap_scripting_security_prompt(info, wrappers):
                    return 0
                continue
            if not is_single_ok_dialog(info):
                continue
            button_index = select_standard_button(info, STANDARD_OK_CONTROL_ID)
            if button_index is None:
                continue
            click_wrapper(wrappers[button_index])
            return 0
        time.sleep(0.35)
    return 0


def close_open_workbook_window(output_path: Path, log_file: Path | None = None) -> bool:
    require_windows_runtime()
    markers = (output_path.name.casefold(), output_path.stem.casefold())
    closed = False
    for window in usable_top_windows():
        title = str(safe_call(window, "window_text", ""))
        if not any(marker in title.casefold() for marker in markers):
            continue
        class_name = str(safe_call(window, "class_name", ""))
        if class_name not in {"XLMAIN", "etMainWindow"}:
            continue
        if title.lstrip().startswith("*"):
            raise RuntimeError(f"output workbook is modified and open: {title}")
        log_line(log_file, f"closing open workbook window: {title}")
        try:
            window.close()
        except Exception as exc:
            raise RuntimeError(f"could not close workbook window: {title} ({exc})") from exc
        deadline = time.time() + 10
        while time.time() < deadline:
            if not any(
                marker in str(safe_call(item, "window_text", "")).casefold()
                for item in usable_top_windows()
                for marker in markers
            ):
                closed = True
                break
            time.sleep(0.3)
        if not closed:
            raise RuntimeError(f"workbook window stayed open: {title}")
    return closed


def wait_for_output_update(path: Path, start_time: float, timeout: int = 300) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        if path.exists() and path.stat().st_mtime >= start_time - 1:
            return
        time.sleep(0.5)
    raise TimeoutError(f"output file was not updated: {path}")


def wait_for_helper_ready(process: subprocess.Popen, ready_file: Path, timeout: int = 20) -> None:
    deadline = time.time() + timeout
    while time.time() < deadline:
        if ready_file.exists():
            return
        result = process.poll()
        if result is not None:
            raise RuntimeError(f"UI helper exited before ready with code {result}")
        time.sleep(0.1)
    process.terminate()
    process.wait(timeout=10)
    raise TimeoutError("UI helper did not become ready")


def remove_ready_file(ready_file: Path) -> None:
    for _ in range(10):
        try:
            ready_file.unlink(missing_ok=True)
            return
        except PermissionError:
            time.sleep(0.1)


def connect_sap_with_access_helper(output_dir: Path):
    require_windows_runtime()
    ready_file = output_dir / (
        f".mb5b-access-ready-{os.getpid()}-{datetime.now():%Y%m%d%H%M%S%f}.tmp"
    )
    process = subprocess.Popen(
        [
            sys.executable,
            str(Path(__file__).resolve()),
            "--access-helper",
            "--ready-file",
            str(ready_file),
        ],
        cwd=str(output_dir),
        creationflags=subprocess.CREATE_NO_WINDOW,
    )
    try:
        wait_for_helper_ready(process, ready_file)
        return get_sap_session()
    finally:
        if process.poll() is None:
            process.terminate()
            process.wait(timeout=10)
        remove_ready_file(ready_file)


def submit_sap_export_as_dialog(session, output_path: Path, log_file: Path) -> bool:
    deadline = time.time() + 20
    while time.time() < deadline:
        if has_sap_control(session, SAP_EXPORT_AS_FILE_ID):
            sap_required(session, SAP_EXPORT_AS_FILE_ID).Text = PureWindowsPath(
                str(output_path)
            ).stem
            sap_required(session, SAP_EXPORT_AS_BUTTON_ID).Press()
            wait_sap(session)
            log_line(log_file, "submitted SAP internal Export As dialog")
            return True
        time.sleep(0.25)
    return False


def submit_sap_save_file_dialog(
    session,
    output_path: Path,
    overwrite: bool,
    log_file: Path,
) -> bool:
    deadline = time.time() + 20
    while time.time() < deadline:
        if has_sap_control(session, SAP_SAVE_FILE_PATH_ID) and has_sap_control(
            session, SAP_SAVE_FILE_NAME_ID
        ):
            sap_required(session, SAP_SAVE_FILE_PATH_ID).Text = str(output_path.parent)
            sap_required(session, SAP_SAVE_FILE_NAME_ID).Text = output_path.name
            button_id = (
                SAP_SAVE_FILE_REPLACE_BUTTON_ID
                if overwrite
                else SAP_SAVE_FILE_GENERATE_BUTTON_ID
            )
            sap_required(session, button_id).Press()
            wait_sap(session)
            log_line(log_file, "submitted SAP internal Save File dialog")
            return True
        time.sleep(0.25)
    return False


def run_one_sap_export(
    session,
    config: RunConfig,
    plant: str,
    storage: str,
    output_path: Path,
    log_file: Path,
) -> None:
    if output_path.exists() and not config.overwrite:
        raise FileExistsError(f"output already exists; pass --overwrite: {output_path}")
    close_open_workbook_window(output_path, log_file)

    sap_required(session, "wnd[0]").Maximize()
    sap_required(session, "wnd[0]/tbar[0]/okcd").Text = "/nmb5b"
    sap_required(session, "wnd[0]").SendVKey(0)
    wait_sap(session)
    if not has_sap_control(session, SAP_PLANT_FIELD_ID):
        time.sleep(1.0)
        sap_required(session, "wnd[0]/tbar[0]/okcd").Text = "/nmb5b"
        sap_required(session, "wnd[0]").SendVKey(0)
        wait_sap(session)
    wait_for_mb5b_selection_screen(session)

    selected_by = select_storage_location_batch_stock(session)
    log_line(log_file, f"selected stock type by technical control {selected_by}")
    sap_required(session, SAP_PLANT_FIELD_ID).Text = plant
    sap_required(session, SAP_STORAGE_FIELD_ID).Text = storage
    sap_required(session, SAP_DATE_FIELD_ID).Text = config.export_date.sap_value
    sap_required(session, SAP_DATE_FIELD_ID).SetFocus()
    sap_required(session, SAP_DATE_FIELD_ID).caretPosition = len(config.export_date.sap_value)

    sap_required(session, SAP_EXECUTE_BUTTON_ID).Press()
    wait_sap(session)
    log_line(log_file, f"after execute: {sap_info(session)}")
    if not has_sap_control(session, SAP_EXPORT_BUTTON_ID):
        raise RuntimeError(
            f"export button unavailable: {sap_info(session)}; status={sap_status_text(session)}"
        )

    export_start = time.time()
    command = [
        sys.executable,
        str(Path(__file__).resolve()),
        "--dialog-helper",
        str(output_path),
    ]
    for process_id in sorted(detect_sap_process_ids()):
        command.extend(["--allowed-pid", str(process_id)])
    ready_file = config.output_dir / (
        f".mb5b-helper-ready-{os.getpid()}-{datetime.now():%Y%m%d%H%M%S%f}.tmp"
    )
    command.extend(["--ready-file", str(ready_file)])
    if config.overwrite:
        command.append("--overwrite")
    helper = subprocess.Popen(
        command,
        cwd=str(config.output_dir),
        creationflags=subprocess.CREATE_NO_WINDOW,
    )
    log_line(log_file, f"started language-independent dialog helper pid={helper.pid}")
    wait_for_helper_ready(helper, ready_file)
    sap_required(session, SAP_EXPORT_BUTTON_ID).Press()
    wait_sap(session)
    submit_sap_export_as_dialog(session, output_path, log_file)
    submit_sap_save_file_dialog(session, output_path, config.overwrite, log_file)

    try:
        helper_result = helper.wait(timeout=320)
    except subprocess.TimeoutExpired as exc:
        helper.terminate()
        helper.wait(timeout=10)
        remove_ready_file(ready_file)
        raise RuntimeError("dialog helper timed out") from exc
    remove_ready_file(ready_file)
    if helper_result != 0:
        raise RuntimeError(f"dialog helper failed with exit code {helper_result}")
    wait_for_output_update(output_path, export_start)
    close_open_workbook_window(output_path, log_file)
    log_line(log_file, f"saved export: {output_path}")


def add_storage_column(workbook_path: Path, storage: str, header: str) -> None:
    workbook = openpyxl.load_workbook(workbook_path)
    try:
        worksheet = (
            workbook[DATA_SHEET]
            if DATA_SHEET in workbook.sheetnames
            else workbook[workbook.sheetnames[0]]
        )
        if worksheet.cell(1, 4).value != header:
            worksheet.insert_cols(4)
        worksheet.cell(1, 4).value = header
        for row_number in range(2, worksheet.max_row + 1):
            worksheet.cell(row_number, 4).value = storage
        workbook.save(workbook_path)
    finally:
        workbook.close()


def copy_cell(source, destination) -> None:
    destination.value = source.value
    if source.has_style:
        destination.font = copy(source.font)
        destination.fill = copy(source.fill)
        destination.border = copy(source.border)
        destination.alignment = copy(source.alignment)
        destination.number_format = source.number_format
        destination.protection = copy(source.protection)


def merge_plant_files(
    output_dir: Path,
    plant: str,
    paths: Sequence[Path],
    date_token: str,
    overwrite: bool,
    log_file: Path | None,
) -> tuple[Path, int]:
    output_path = plant_merge_output_path(output_dir, plant, date_token)
    if output_path.exists() and not overwrite:
        raise FileExistsError(f"merge output already exists; pass --overwrite: {output_path}")

    output_workbook = openpyxl.Workbook()
    output_worksheet = output_workbook.active
    output_worksheet.title = DATA_SHEET
    next_row = 1
    header_written = False
    total_rows = 0

    try:
        for path in paths:
            source_workbook = openpyxl.load_workbook(path, data_only=False)
            try:
                source_worksheet = (
                    source_workbook[DATA_SHEET]
                    if DATA_SHEET in source_workbook.sheetnames
                    else source_workbook[source_workbook.sheetnames[0]]
                )
                start_row = 1 if not header_written else 2
                for row_number in range(start_row, source_worksheet.max_row + 1):
                    for column_number in range(1, source_worksheet.max_column + 1):
                        copy_cell(
                            source_worksheet.cell(row_number, column_number),
                            output_worksheet.cell(next_row, column_number),
                        )
                    next_row += 1
                    if row_number >= 2:
                        total_rows += 1
                header_written = True
            finally:
                source_workbook.close()

        for column_cells in output_worksheet.columns:
            maximum_length = 0
            column_letter = column_cells[0].column_letter
            for cell in column_cells[:200]:
                value = "" if cell.value is None else str(cell.value)
                maximum_length = max(maximum_length, len(value))
            output_worksheet.column_dimensions[column_letter].width = min(
                max(maximum_length + 2, 8), 30
            )
        output_workbook.save(output_path)
    finally:
        output_workbook.close()

    log_line(
        log_file,
        f"merged plant={plant}, files={len(paths)}, rows={total_rows}, output={output_path}",
    )
    return output_path, total_rows


def build_config(args: argparse.Namespace) -> RunConfig:
    if args.input is None:
        raise ValueError("--input is required")
    if args.date is None:
        raise ValueError("--date is required")
    input_path = args.input.expanduser().resolve()
    output_dir = (
        args.output_dir.expanduser().resolve()
        if args.output_dir is not None
        else input_path.parent
    )
    output_dir.mkdir(parents=True, exist_ok=True)
    return RunConfig(
        input_path=input_path,
        export_date=args.date,
        output_dir=output_dir,
        sheet=args.sheet,
        plant_column=args.plant_column,
        storage_column=args.storage_column,
        storage_header=args.storage_header,
        limit=max(0, args.limit),
        dry_run=args.dry_run,
        overwrite=args.overwrite,
    )


def run_main(args: argparse.Namespace) -> int:
    try:
        config = build_config(args)
        targets = read_targets(
            config.input_path,
            config.sheet,
            config.plant_column,
            config.storage_column,
            config.limit,
        )
        if not targets:
            raise ValueError("input workbook contains no valid plant/storage rows")
    except Exception as exc:
        print(f"ERROR: {type(exc).__name__}: {exc}", file=sys.stderr)
        return 1

    log_file = config.output_dir / f"MB5B_export_log_{datetime.now():%Y%m%d_%H%M%S}.txt"
    log_line(log_file, "start sap-mb5b-export")
    log_line(log_file, f"input={config.input_path}")
    log_line(log_file, f"output_dir={config.output_dir}")
    log_line(
        log_file,
        f"export_date={config.export_date.sap_value} token={config.export_date.token}",
    )
    log_line(
        log_file,
        f"target_count={len(targets)} dry_run={config.dry_run} overwrite={config.overwrite}",
    )

    planned_groups: dict[str, list[Path]] = {}
    for target in targets:
        planned_groups.setdefault(target["plant"], []).append(
            individual_output_path(
                config.output_dir,
                target["plant"],
                target["storage"],
                config.export_date.token,
            )
        )

    if config.dry_run:
        for index, target in enumerate(targets, start=1):
            path = individual_output_path(
                config.output_dir,
                target["plant"],
                target["storage"],
                config.export_date.token,
            )
            log_line(
                log_file,
                f"DRY target {index}: plant={target['plant']}, storage={target['storage']}, output={path}",
            )
        for plant, paths in planned_groups.items():
            log_line(
                log_file,
                f"DRY merge: plant={plant}, files={len(paths)}, "
                f"output={plant_merge_output_path(config.output_dir, plant, config.export_date.token)}",
            )
        return 0

    try:
        require_windows_runtime()
        session = connect_sap_with_access_helper(config.output_dir)
        log_line(log_file, f"SAP connected: {sap_info(session)}")
    except Exception as exc:
        log_line(log_file, f"FATAL SAP connection error: {type(exc).__name__}: {exc}")
        return 1

    failures: list[tuple[str, str, str]] = []
    merge_groups: dict[str, list[Path]] = {}
    for index, target in enumerate(targets, start=1):
        plant = target["plant"]
        storage = target["storage"]
        output_path = individual_output_path(
            config.output_dir, plant, storage, config.export_date.token
        )
        log_line(
            log_file,
            f"START {index}/{len(targets)} plant={plant}, storage={storage}, output={output_path}",
        )
        try:
            run_one_sap_export(session, config, plant, storage, output_path, log_file)
            add_storage_column(output_path, storage, config.storage_header)
            log_line(log_file, f"added storage column: {output_path}")
            merge_groups.setdefault(plant, []).append(output_path)
        except Exception as exc:
            message = f"{type(exc).__name__}: {exc}"
            failures.append((plant, storage, message))
            log_line(log_file, f"ERROR plant={plant}, storage={storage}: {message}")

    for plant, paths in merge_groups.items():
        try:
            merge_plant_files(
                config.output_dir,
                plant,
                paths,
                config.export_date.token,
                config.overwrite,
                log_file,
            )
        except Exception as exc:
            message = f"{type(exc).__name__}: {exc}"
            failures.append((plant, "<merge>", message))
            log_line(log_file, f"ERROR plant={plant}, merge: {message}")

    log_line(
        log_file,
        f"finished successes={sum(len(paths) for paths in merge_groups.values())} "
        f"failures={len(failures)} log={log_file}",
    )
    return 2 if failures else 0


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(
        description="Language-independent SAP GUI automation for MB5B Excel exports"
    )
    parser.add_argument("--input", type=Path)
    parser.add_argument("--date", type=parse_export_date)
    parser.add_argument("--output-dir", type=Path)
    parser.add_argument("--sheet", default=DEFAULT_INPUT_SHEET)
    parser.add_argument("--plant-column", default=DEFAULT_PLANT_HEADER)
    parser.add_argument("--storage-column", default=DEFAULT_STORAGE_HEADER)
    parser.add_argument("--storage-header", default=DEFAULT_STORAGE_HEADER)
    parser.add_argument("--limit", type=int, default=0)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--overwrite", action="store_true")
    parser.add_argument("--inspect-ui", action="store_true")
    parser.add_argument("--dialog-helper", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--access-helper", action="store_true", help=argparse.SUPPRESS)
    parser.add_argument("--ready-file", type=Path, help=argparse.SUPPRESS)
    parser.add_argument("--allowed-pid", type=int, action="append", help=argparse.SUPPRESS)
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    parser = build_parser()
    args = parser.parse_args(argv)
    if args.access_helper:
        return access_helper(ready_file=args.ready_file)
    if args.dialog_helper:
        return dialog_helper(
            args.dialog_helper.expanduser().resolve(),
            args.overwrite,
            ready_file=args.ready_file,
            allowed_process_ids=set(args.allowed_pid or []),
        )
    if args.inspect_ui:
        output_dir = (
            args.output_dir.expanduser().resolve()
            if args.output_dir is not None
            else Path.cwd()
        )
        output_dir.mkdir(parents=True, exist_ok=True)
        return inspect_ui(output_dir)
    return run_main(args)


if __name__ == "__main__":
    raise SystemExit(main())
